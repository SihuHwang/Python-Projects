import openpyxl

wb = openpyxl.load_workbook('openpyxltest.xlsx')

#ws = wb.active

ws = wb['Sheet']

cell1 = ws['C4'].value

print(cell1)

cell2 = ws.cell(row=2, column=3).value

print(cell2)

print(ws.rows)

ws['A1'] = 'Test1'
ws.cell(row=2, column=1).value = 'Test2'

for r in  ws.rows:
    print(r)
    print(r[0].value)
    print(r[1].value)
    print(r[2].value)
    print(r[3].value)
    print('#############')

for c in ws.columns:
    print(c)
    print(c[0].value)
    print(c[1].value)
    print(c[2].value)
    print(c[3].value)
    print('#############')


cell_range = ws['B1':'D6']
print(cell_range)
for col_cell in cell_range:
    for cell in col_cell:
        print(cell.value, end = " , ")
    print('')

row3 = ws[3]
print(row3)
for col_cell in row3:
    print(col_cell.value)

row_range = ws[1:6]
print(row_range)
for col_cell in row_range:
    for cell in col_cell:
        print(cell.value, end = " , ")
    print('')

colC = ws['C']
print(colC)
for col_cell in colC:
    print(col_cell.value)

col_range = ws['A:D']
print(col_range)
for col_cell in col_range:
    for cell in col_cell:
        print(cell.value, end= ' , ')
    print('')

ws.append(['a','b','c'])



wb.save('openpyxltest.xlsx')
wb.close()

 